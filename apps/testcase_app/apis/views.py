from rest_framework.views import Response
from rest_framework import generics
from apps.testcase_app.models import TestCaseModel, TestCaseStep, NatcoStatus, TestResult
from apps.testcase_app.apis.serializers import TestCaseSerializerList, TestCaseSerializer, ExcelSerializer, \
        NatcoStatusSerializer, TestCaseStatusUpdateSerializer, TestResultDRPSerializer, DistinctTestResultSerializer
from apps.stbs.models import NactoManufactureLanguage
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from apps.testcase_app.pagination import CustomPagination
from openpyxl import load_workbook
from django.db import transaction
from django.forms.models import model_to_dict
from django.http import JsonResponse
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiSchemaBase
from drf_spectacular.openapi import OpenApiTypes, OpenApiExample
from rest_framework.exceptions import APIException
from django_filters import rest_framework as filters
from apps.testcase_app.filters import NatcoStatusFilter
from rest_framework import status
from apps.stbs.permissions import AdminPermission
from qa_backend.helpers.renders import ResponseInfo
from rest_framework.renderers import TemplateHTMLRenderer
from qa_backend.helpers import custom_generics as cgenerics
from django.db.models import OuterRef, Subquery
from django.views.generic import TemplateView
from rest_framework.views import APIView
from rest_framework import status
from django.db.models import Min
import json


class TestCaseStatusUpdateView(generics.GenericAPIView):

    def __init__(self, **kwargs) -> None:
        self.response_format = ResponseInfo().response
        super().__init__(**kwargs)

    serializer_class = TestCaseStatusUpdateSerializer
    queryset = TestCaseModel.objects.all()

    def patch(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.update(serializer.validated_data)
            self.response_format['status'] = True
            self.response_format['status_code'] = status.HTTP_200_OK
            self.response_format['data'] = "Success"
            self.response_format['message'] = "Success"
            return Response(self.response_format, status=status.HTTP_200_OK)
        else:
            self.response_format['status'] = False
            self.response_format['status_code'] = status.HTTP_400_BAD_REQUEST
            self.response_format['message'] = "error"
            return Response(self.response_format, status=status.HTTP_400_BAD_REQUEST)


class TestCaseListView(generics.ListAPIView):

    # authentication_classes = [JWTAuthentication]
    # permission_classes = [AdminPermission]
    queryset = TestCaseModel.objects.all()
    serializer_class = TestCaseSerializerList
    pagination_class = CustomPagination
    filter_backends = [filters.DjangoFilterBackend]
    filterset_fields = ('jira_id', 'test_name', 'status', 'automation_status')

    def get(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        return super().list(request, *args, **kwargs)


class TestCaseView(cgenerics.CustomCreateAPIView):

    # permission_classes = [AdminPermission]
    serializer_class = TestCaseSerializer

    def post(self, request, *args, **kwargs):
        return super(TestCaseView, self).post(request, *args, **kwargs)


class TestCaseDetailView(cgenerics.CustomRetrieveUpdateDestroyAPIView):

    # permission_classes = [AdminPermission]
    lookup_field = 'jira_id'
    serializer_class = TestCaseSerializer
    
    def get_object(self):
        queryset = get_object_or_404(TestCaseModel.objects.prefetch_related('test_steps'), jira_id=self.kwargs.get('jira_id'))
        # natco = queryset.annotate(natco_status=Subquery(NatcoStatus.objects.select_related('test_case', 'language', 'device', 'natco', 'user').filter(test_case_id=self.kwargs.get('jira_id'))))
        return queryset

    def get(self, request, *args, **kwargs):
        response = super().retrieve(request, *args, **kwargs)
        queryset = NatcoStatus.objects.select_related('test_case', 'language', 'device', 'natco', 'user').filter(test_case_id=self.kwargs.get('jira_id'))
        serializer = NatcoStatusSerializer(queryset, many=True)
        response.data['natco_status'] = serializer.data
        return response

class TestCaseNatcoView(generics.ListAPIView):

    # permission_classes = [AdminPermission]
    serializer_class = NatcoStatusSerializer
    queryset = NatcoStatus.objects.all()
    lookup_field = 'jira_id'
    pagination_class = CustomPagination

    def get_queryset(self):
        queryset = NatcoStatus.objects.select_related('test_case', 'natco_status').filter(test_case_id=self.kwargs.get('jira_id')).values('jira_id', 'jira_summary')
        return queryset


class TestCaseNatcoList(generics.ListAPIView):
    # permission_classes = [AdminPermission]
    serializer_class = NatcoStatusSerializer
    filterset_class = NatcoStatusFilter
    pagination_class = CustomPagination

    def get_queryset(self):
        queryset = NatcoStatus.objects.select_related('test_case', 'language', 'device', 'natco', 'user').all()
        return queryset

    @extend_schema(
        parameters=[
            OpenApiParameter(name='Natco', description="Enter the Natco", required=False, type=OpenApiTypes.STR,
                             location=OpenApiParameter.QUERY),
            OpenApiParameter(name='Language', description="Enter the Language", required=False, type=OpenApiTypes.STR,
                             location=OpenApiParameter.QUERY),
            OpenApiParameter(name='Device', description="Enter the Device", required=False, type=OpenApiTypes.STR,
                             location=OpenApiParameter.QUERY),
            OpenApiParameter(name='Jira Id', description="Enter the Jira ID", required=False, type=OpenApiTypes.STR,
                             location=OpenApiParameter.QUERY),
            OpenApiParameter(name='Applicable', description="Enter the Applicable", required=False, type=OpenApiTypes.BOOL,
                             location=OpenApiParameter.QUERY),
            OpenApiParameter(name='status', description="Choose a Status", required=False,
                             type=OpenApiTypes.STR,
                             location=OpenApiParameter.QUERY, enum=NatcoStatus.STATUS_CHOICES),
        ]
    )
    def list(self, request, *args, **kwargs):
        data = self.get_queryset()
        filter_set = self.filterset_class(request.GET, self.get_queryset())
        if filter_set.is_valid():
            data = filter_set.qs
        paginated_data = self.paginate_queryset(data)
        serializer = self.get_serializer(paginated_data, many=True)
        try:
            if serializer:
                return self.get_paginated_response(serializer.data)
        except Exception as e:
            return Response({"success": False, "data": str(e)})
    

class TestCaseNatcoDetail(cgenerics.CustomRetrieveUpdateDestroyAPIView):
    # permission_classes = [AdminPermission]
    serializer_class = NatcoStatusSerializer
    # queryset = NatcoStatus.objects.all()
    lookup_field = 'pk'

    def get_object(self):
        queryset = NatcoStatus.objects.get(id=self.kwargs.get('pk')).select_related('test_case')
        return queryset
    

class TestResultFilterView(generics.GenericAPIView):

    def __init__(self, **kwargs):
        self.response_format = ResponseInfo().response
        super().__init__(**kwargs)

    serializer_class = TestResultDRPSerializer

    def get_queryset(self):
        queryset = TestResult.get_unique_filters()
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if queryset:
            self.response_format['success'] = True
            self.response_format['status_code'] = status.HTTP_200_OK
            self.response_format['data'] = queryset
            self.response_format['message'] = 'Success'
            return Response(self.response_format, status=status.HTTP_200_OK)
        if not queryset:
            self.response_format['success'] = False
            self.response_format['status_code'] = status.HTTP_400_BAD_REQUEST
            self.response_format['message'] = 'Error'
            return Response(self.response_format, status=status.HTTP_200_OK)
        return Response(self.response_format, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class TestCaseReportView(generics.GenericAPIView):

    def __init__(self, **kwargs):
        self.response_format = ResponseInfo().response
        super().__init__(**kwargs)

    serializer_class = DistinctTestResultSerializer
    
    def get_queryset(self):
        results = TestResult.objects.values('testcase', 'natco').annotate(min_cpu=Min('cpu_usage'), min_ram=Min('ram_usage'), min_time=Min('load_time'))

        # Now, let's filter the results to get the distinct testcase and distinct natco with the minimum values
        distinct_results = []
        for result in results:
            distinct_result = TestResult.objects.filter(testcase=result['testcase'], natco=result['natco'], cpu_usage=result['min_cpu'], ram_usage=result['min_ram'], load_time=result['min_time']).first()
            if distinct_result:
                distinct_results.append(distinct_result)
        return distinct_results
    
    def get(self, request, *args, **kwargs):
        serializer = self.get_serializer(self.get_queryset(), many=True)
        if serializer.data:
            self.response_format['success'] = True
            self.response_format['status_code'] = status.HTTP_200_OK
            self.response_format['data'] = serializer.data
            self.response_format['message'] = 'Success'
            return Response(self.response_format, status=status.HTTP_200_OK)
        if not serializer.data:
            self.response_format['success'] = False
            self.response_format['status_code'] = status.HTTP_400_BAD_REQUEST
            self.response_format['message'] = 'Error'
            return Response(self.response_format, status=status.HTTP_200_OK)
        return Response(self.response_format, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GetTestResult(generics.GenericAPIView):

    def __init__(self, **kwargs):
        self.response_format = ResponseInfo().response
        super().__init__(**kwargs)

    serializer_class = ExcelSerializer

    def post(self, request, *args, **kwargs):
        _file = request.FILES.get('file')
        wb = load_workbook(_file)
        ws = wb.active
        _test_result = []
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                _data = {
                    'run_type': row[0], 
                    'date': row[1], 
                    'iteration_number' : row[2], 
                    'testcase' : row[3], 
                    'load_time' : row[4], 
                    'cpu' : row[5], 
                    'ram' : row[6], 
                    'start_time' : row[7], 
                    'end_time' : row[8], 
                    'job_uid' : row[9], 
                    'node_id' : row[10], 
                    'failure_reason' : row[11], 
                    'result' : row[12], 
                    'natco' : row[13], 
                    'load_time' : row[14], 
                    'cpu_usage' : row[15], 
                    'ram_usage' : row[16], 
                    'country_code' : row[17], 
                    'stb_release' : row[18], 
                    'stb_firmware' : row[19], 
                    'stb_android' : row[20], 
                    'stb_build' : row[21], 
                    'natoc_node' : row[22], 
                    'comment' : row[23], 
                }
                _test_result.append(TestResult(**_data))
            with transaction.atomic():
                TestResult.objects.bulk_create(_test_result)
        except Exception as e:
            self.response_format['status'] = False
            self.response_format['status_code'] = status.HTTP_400_BAD_REQUEST
            self.response_format['data'] = 'Error'
            self.response_format['massage'] = "mesga"
            return Response(self.response_format, status=status.HTTP_400_BAD_REQUEST)
        self.response_format['status'] = True
        self.response_format['status_code'] = status.HTTP_200_OK
        self.response_format['data'] = "Success"
        self.response_format['message'] = "TestCase Uploaded Successfully"
        return Response(self.response_format, status=status.HTTP_201_CREATED)
        


class GetTestCase(generics.GenericAPIView):

    def __init__(self, **kwargs) -> None:
        self.response_format = ResponseInfo().response
        super().__init__(**kwargs)

    # permission_classes = [AdminPermission]
    serializer_class = ExcelSerializer

    def post(self, request, *args, **kwargs):
        file_uploaded = request.FILES.get("file")
        wb = load_workbook(file_uploaded)
        ws = wb.active
        test_case = None
        _data = dict()
        _step_data = dict()
        testcase_list = []
        step_list = []
        natco_list = []
        natco = NactoManufactureLanguage.objects.all()
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] is not None:
                    jira_id_parts = str(row[2]).split('-')
                    _data = {
                        "jira_id": jira_id_parts[-1],
                        "jira_summary": row[6],
                        "test_description": row[7],
                        "test_name": f"TestCase - {jira_id_parts[-1]}"
                    }
                    testcase_list.append(TestCaseModel(**_data))
                    test_case = jira_id_parts[-1]
                    for data in natco:
                        natco_list.append(NatcoStatus(natco=data.natco, language=data.language_name,
                                                      device=data.device_name, test_case_id=test_case))
                    if row[2] and row[8]:
                        _step_data = {
                        "testcase_id": test_case,
                        "step_id": int(row[8]),
                        "step_description": row[9],
                        "step_data": '',
                        "excepted_result": row[10]
                        }
                    step_list.append(TestCaseStep(**_step_data))
                elif row[2] is None and row[8] is not None:
                    _step_data = {
                        "testcase_id": test_case,
                        "step_id": int(row[8]),
                        "step_description": row[9],
                        "step_data": '',
                        "excepted_result": row[10]
                    }
                    step_list.append(TestCaseStep(**_step_data))
                elif row[2] is None and row[8] is None:
                    pass
            with transaction.atomic():
                TestCaseModel.objects.bulk_create(testcase_list)
                TestCaseStep.objects.bulk_create(step_list)
                NatcoStatus.objects.bulk_create(natco_list)
        except Exception as e:
            self.response_format['status'] = False
            self.response_format['status_code'] = status.HTTP_400_BAD_REQUEST
            self.response_format['data'] = 'Error'
            self.response_format['massage'] = "TestCase Upload Failed"
            return Response(self.response_format, status=status.HTTP_400_BAD_REQUEST)
        self.response_format['status'] = True
        self.response_format['status_code'] = status.HTTP_200_OK
        self.response_format['data'] = "Success"
        self.response_format['message'] = "TestCase Uploaded Successfully"
        return Response(self.response_format, status=status.HTTP_201_CREATED)


